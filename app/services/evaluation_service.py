from typing import List, Dict, Any, Optional, Tuple
from app import db
from app.models import (
    ModelEvaluation, 
    ModelEvaluationDataset, 
    ModelEvaluationResult, 
    AIModel, 
    Dataset, 
)
from flask import current_app
import threading
from app.services.model_service import get_decrypted_api_key
from app.utils import get_beijing_time
from collections import OrderedDict, defaultdict
from evalscope.run import run_task
from evalscope.constants import JudgeStrategy
import os
import json
import pandas as pd

# å¯¼å…¥é…ç½®å‡½æ•°
from app.config import get_outputs_dir

# è¾…åŠ©å‡½æ•°ï¼Œå°è¯•å°†evalscopeçš„Reportå¯¹è±¡è½¬æ¢ä¸ºå¯åºåˆ—åŒ–çš„å­—å…¸
def serialize_evalscope_report(report_obj):
    if hasattr(report_obj, 'to_dict') and callable(report_obj.to_dict):
        return report_obj.to_dict()
    try:
        return vars(report_obj) 
    except TypeError:
        return str(report_obj) 

# å®šä¹‰evalscopeè¾“å‡ºç»“æ„ä¸­reviewç›®å½•çš„åç§°
OUTPUTS_STRUCTURE_REVIEWS_DIR = 'reviews'

# ç®€å•çš„ç¼“å­˜æœºåˆ¶ï¼Œå­˜å‚¨è¯„ä¼°ä»»åŠ¡çš„total_prompts
_evaluation_total_prompts_cache = {}

class EvaluationService:
    """æ¨¡å‹è¯„ä¼°æœåŠ¡ï¼Œå¤„ç†è¯„ä¼°ç›¸å…³çš„ä¸šåŠ¡é€»è¾‘"""
    
    @staticmethod
    def create_evaluation(
        user_id: int, 
        model_id: int, 
        judge_model_id: int,
        datasets: List[Dict[str, Any]],
        temperature: float, 
        max_tokens: int,
        top_k: Optional[int] = None,  # æ–°å¢top_kå‚æ•°
        top_p: Optional[float] = None,  # æ–°å¢top_på‚æ•°
        name: Optional[str] = None,
        limit: Optional[int] = None,  # æ–°å¢ limit å‚æ•°
        judge_worker_num: Optional[int] = None,  # æ–°å¢å¹¶å‘æ•°å‚æ•°
        eval_batch_size: Optional[int] = None  # æ–°å¢è¯„ä¼°å¹¶å‘æ•°å‚æ•°
    ) -> Optional[ModelEvaluation]:
        """
        åˆ›å»ºä¸€ä¸ªæ–°çš„æ¨¡å‹è¯„ä¼°ä»»åŠ¡
        """
        try:
            if not name:
                model = AIModel.query.get(model_id)
                name = f"{model.display_name if model else 'æœªçŸ¥æ¨¡å‹'}çš„è¯„ä¼°_{get_beijing_time().strftime('%Y%m%d_%H%M%S')}"
            
            evaluation = ModelEvaluation(
                user_id=user_id,
                model_id=model_id,
                judge_model_id=judge_model_id,
                name=name,
                temperature=temperature,
                max_tokens=max_tokens,
                top_k=top_k,  # æ·»åŠ top_k
                top_p=top_p,  # æ·»åŠ top_p
                judge_worker_num=judge_worker_num,  # æ·»åŠ å¹¶å‘æ•°
                eval_batch_size=eval_batch_size,  # æ·»åŠ è¯„ä¼°å¹¶å‘æ•°
                status='pending',
                limit=limit  # ä¿å­˜ limit å€¼
            )
            db.session.add(evaluation)
            db.session.flush()
            
            for dataset_info in datasets:
                dataset_id = dataset_info.get('dataset_id')
                eval_dataset = ModelEvaluationDataset(
                    evaluation_id=evaluation.id,
                    dataset_id=dataset_id,
                    subset=None, 
                    split=None   
                )
                db.session.add(eval_dataset)
            
            db.session.commit()
            
            app_context = current_app._get_current_object()
            threading.Thread(
                target=EvaluationService._run_evaluation_task, 
                args=(app_context, evaluation.id,)
            ).start()
            
            return evaluation
        
        except Exception as e:
            current_app.logger.error(f"åˆ›å»ºè¯„ä¼°ä»»åŠ¡å¤±è´¥: {str(e)}", exc_info=True)
            db.session.rollback()
            return None
    
    @staticmethod
    def _run_evaluation_task(app, evaluation_id: int) -> None: 
        with app.app_context(): 
            current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] å¼€å§‹æ‰§è¡Œã€‚")
            evaluation = ModelEvaluation.query.get(evaluation_id)
            if not evaluation:
                current_app.logger.error(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] æ— æ³•æ‰¾åˆ°è¯„ä¼°è®°å½•ã€‚")
                return
            
            evaluation.status = 'running'
            db.session.commit()
            current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] çŠ¶æ€æ›´æ–°ä¸º 'running'ã€‚")
            
            model_to_evaluate = AIModel.query.get(evaluation.model_id)
            judge_model_for_evalscope = None if evaluation.judge_model_id is None else AIModel.query.get(evaluation.judge_model_id)

            if not model_to_evaluate:
                evaluation.status = 'failed'
                evaluation.result_summary = {"error": "è¢«è¯„ä¼°æ¨¡å‹ä¸å­˜åœ¨"}
                db.session.commit()
                current_app.logger.error(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] å¤±è´¥: è¢«è¯„ä¼°æ¨¡å‹ ID {evaluation.model_id} ä¸å­˜åœ¨ã€‚")
                return
            
            judge_api_url = None
            judge_api_key = None
            judge_model_identifier = None
            if judge_model_for_evalscope:
                judge_api_url = judge_model_for_evalscope.api_base_url
                judge_api_key = get_decrypted_api_key(judge_model_for_evalscope)
                judge_model_identifier = judge_model_for_evalscope.model_identifier

            eval_dataset_associations = ModelEvaluationDataset.query.filter_by(evaluation_id=evaluation_id).all()
            dataset_names_for_evalscope = []
            dataset_args = {}  # æ–°å¢ï¼šä¸ºè‡ªå»ºæ•°æ®é›†å‡†å¤‡çš„dataset_args

            # è·å–æ‰€æœ‰å‚ä¸è¯„ä¼°çš„æ•°æ®é›†çš„åç§° (è¿™äº›æ˜¯ä¼ é€’ç»™evalscopeçš„åç§°)
            for assoc in eval_dataset_associations:
                dataset = Dataset.query.get(assoc.dataset_id)
                if dataset:
                    if dataset.dataset_type == 'ç³»ç»Ÿ':
                        # ç³»ç»Ÿæ•°æ®é›†ç›´æ¥ä½¿ç”¨åç§°
                        dataset_names_for_evalscope.append(dataset.name) 
                    elif dataset.dataset_type == 'è‡ªå»º':
                        # è‡ªå»ºæ•°æ®é›†æ ¹æ®æ ¼å¼ä½¿ç”¨general_mcqæˆ–general_qa
                        if dataset.format == 'MCQ':
                            if 'general_mcq' not in dataset_names_for_evalscope:
                                dataset_names_for_evalscope.append('general_mcq')
                            
                            # è·å–ä¸Šä¼ ç›®å½•çš„è·¯å¾„
                            dataset_file_path = dataset.download_url
                            dataset_dir = os.path.dirname(dataset_file_path)
                            dataset_name = os.path.splitext(os.path.basename(dataset_file_path))[0]
                            
                            # ç¡®ä¿æœ‰general_mcqçš„dataset_args
                            if 'general_mcq' not in dataset_args:
                                dataset_args['general_mcq'] = {
                                    "local_path": dataset_dir,
                                    "subset_list": [dataset_name],
                                    'filters': {'remove_until': '</think>'} 
                                }
                            else:
                                # å¦‚æœå·²å­˜åœ¨ï¼Œæ·»åŠ åˆ°subset_list
                                if dataset_name not in dataset_args['general_mcq']['subset_list']:
                                    dataset_args['general_mcq']['subset_list'].append(dataset_name)
                            
                        elif dataset.format == 'QA':
                            if 'general_qa' not in dataset_names_for_evalscope:
                                dataset_names_for_evalscope.append('general_qa')
                            
                            # è·å–ä¸Šä¼ ç›®å½•çš„è·¯å¾„
                            dataset_file_path = dataset.download_url
                            dataset_dir = os.path.dirname(dataset_file_path)
                            dataset_name = os.path.splitext(os.path.basename(dataset_file_path))[0]
                            
                            # ç¡®ä¿æœ‰general_qaçš„dataset_args
                            if 'general_qa' not in dataset_args:
                                dataset_args['general_qa'] = {
                                    "local_path": dataset_dir,
                                    "subset_list": [dataset_name],
                                    'filters': {'remove_until': '</think>'} 
                                }
                            else:
                                # å¦‚æœå·²å­˜åœ¨ï¼Œæ·»åŠ åˆ°subset_list
                                if dataset_name not in dataset_args['general_qa']['subset_list']:
                                    dataset_args['general_qa']['subset_list'].append(dataset_name)
                        
                        elif dataset.format == 'CUSTOM':
                            # è·å–ä¸Šä¼ ç›®å½•çš„è·¯å¾„
                            dataset_file_path = dataset.download_url
                            dataset_dir = os.path.dirname(dataset_file_path)
                            dataset_name = os.path.splitext(os.path.basename(dataset_file_path))[0]
                            
                            # åŠ¨æ€æ³¨å†Œè‡ªå®šä¹‰æ•°æ®é›†åŸºå‡†æµ‹è¯•
                            from app.adapter.custom_dataset_adapter import register_custom_dataset_benchmark
                            custom_dataset_key = register_custom_dataset_benchmark(dataset.id)
                            
                            # ä¸ºæ¯ä¸ªè‡ªå®šä¹‰æ•°æ®é›†åˆ›å»ºå•ç‹¬çš„é…ç½®
                            dataset_args[custom_dataset_key] = {
                                "local_path": dataset_dir,
                                "subset_list": [dataset_name],
                                'filters': {'remove_until': '</think>'}
                            }
                            
                            # æ·»åŠ æ¨¡æ¿å†…å®¹
                            if dataset.jinja2_template:
                                dataset_args[custom_dataset_key]['template_content'] = dataset.jinja2_template
                            
                            # æ·»åŠ åˆ°è¯„ä¼°æ•°æ®é›†åˆ—è¡¨
                            dataset_names_for_evalscope.append(custom_dataset_key)
                        current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] æ·»åŠ è‡ªå»ºæ•°æ®é›† {dataset.name}ï¼Œæ ¼å¼: {dataset.format}ï¼Œæ–‡ä»¶è·¯å¾„: {dataset.download_url}")
                else:
                    current_app.logger.warning(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] æ•°æ®é›†ID {assoc.dataset_id} æ— æ³•æ‰¾åˆ°æˆ–åç§°ä¸ºç©ºï¼Œå·²è·³è¿‡ã€‚")
            
            if not dataset_names_for_evalscope:
                evaluation.status = 'failed'
                evaluation.result_summary = {"error": "æ²¡æœ‰æœ‰æ•ˆçš„è¯„ä¼°æ•°æ®é›†"}
                db.session.commit()
                current_app.logger.error(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] å¤±è´¥: æ²¡æœ‰æä¾›æœ‰æ•ˆçš„æ•°æ®é›†è¿›è¡Œè¯„ä¼°ã€‚")
                return

            evalscope_run_timestamp = get_beijing_time().strftime('%Y%m%d_%H%M%S')
            base_output_dir = os.path.join(get_outputs_dir(), f'eval_{evaluation_id}_{evalscope_run_timestamp}') 
            try:
                os.makedirs(base_output_dir, exist_ok=True)
            except Exception as e:
                current_app.logger.error(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] åˆ›å»ºevalscopeè¾“å‡ºç›®å½•å¤±è´¥: {base_output_dir}, error: {e}")
                evaluation.status = 'failed'
                evaluation.result_summary = {"error": f"åˆ›å»ºè¾“å‡ºç›®å½•å¤±è´¥: {e}"}
                db.session.commit()
                return

            decrypted_api_key = get_decrypted_api_key(model_to_evaluate)

            # ä½¿ç”¨TaskConfigæ ¼å¼åˆ›å»ºä»»åŠ¡é…ç½®
            try:
                from evalscope import TaskConfig
            
                task_cfg_args = {
                    'eval_type': 'service', 
                    'api_url': model_to_evaluate.api_base_url,
                    'model': model_to_evaluate.model_identifier, 
                    'api_key': decrypted_api_key if decrypted_api_key else 'NO_API_KEY',
                    'datasets': dataset_names_for_evalscope, # ä¼ é€’é¡¶å±‚æ•°æ®é›†åç§°
                    'stream': True, 
                    'timeout': 12000, 
                    'work_dir': base_output_dir,
                    'generation_config': {
                        'max_new_tokens': evaluation.max_tokens,
                        'temperature': evaluation.temperature,
                        'top_k': evaluation.top_k,
                        'top_p': evaluation.top_p
                    },
                    'eval_batch_size': evaluation.eval_batch_size or 4  # ä½¿ç”¨è¯„ä¼°å¹¶å‘æ•°
                }
                if judge_model_identifier:
                    task_cfg_args['judge_strategy'] = JudgeStrategy.AUTO
                    task_cfg_args['judge_worker_num'] = evaluation.judge_worker_num
                    task_cfg_args['judge_model_args'] = {
                        'model_id': judge_model_identifier,
                        'api_url': judge_api_url if judge_api_url else '',
                        'api_key': judge_api_key if judge_api_key else '',
                        'generation_config': {
                            # 'stream': True,
                            'timeout': 12000,
                        }
                    }
                
                # å¦‚æœæœ‰è‡ªå»ºæ•°æ®é›†ï¼Œæ·»åŠ dataset_argså‚æ•°
                if dataset_args:
                    task_cfg_args['dataset_args'] = dataset_args
                    current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] æ·»åŠ dataset_args: {json.dumps(dataset_args, indent=2)}")
                    
                if evaluation.limit and int(evaluation.limit) > 0:
                    task_cfg_args['limit'] = int(evaluation.limit)

                # ä½¿ç”¨TaskConfigåˆ›å»ºé…ç½®å¯¹è±¡
                task_cfg = TaskConfig(**task_cfg_args)
                
                current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Evalscope task_cfg: {task_cfg_args}")
            except ImportError:
                # å¦‚æœæ— æ³•å¯¼å…¥TaskConfigï¼Œåˆ™å›é€€åˆ°ä½¿ç”¨å­—å…¸
                task_cfg = {
                    'eval_type': 'service', 
                    'api_url': model_to_evaluate.api_base_url,
                    'model': model_to_evaluate.model_identifier, 
                    'api_key': decrypted_api_key if decrypted_api_key else 'NO_API_KEY',
                    'datasets': dataset_names_for_evalscope, # ä¼ é€’é¡¶å±‚æ•°æ®é›†åç§°
                    'stream': True, 
                    'timeout': 12000, 
                    'work_dir': base_output_dir,
                    'generation_config': {
                        'temperature': evaluation.temperature,
                        'max_tokens': evaluation.max_tokens,
                    },
                    'judge_model_args': { 
                        'model_id': judge_model_identifier if judge_model_identifier else '',
                        'api_url': judge_api_url if judge_api_url else '',
                        'api_key': judge_api_key if judge_api_key else ''
                    }
                }
                
                # å¦‚æœæœ‰è‡ªå»ºæ•°æ®é›†ï¼Œæ·»åŠ dataset_argså‚æ•°
                if dataset_args:
                    task_cfg['dataset_args'] = dataset_args
                    current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] æ·»åŠ dataset_args: {json.dumps(dataset_args, indent=2)}")
                    
                if evaluation.limit and int(evaluation.limit) > 0:
                    task_cfg['limit'] = int(evaluation.limit)

                current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Evalscope task_cfg: {json.dumps(task_cfg, indent=2)}")

            evalscope_final_report = {}
            detailed_results_to_save = [] # ç”¨äºå­˜å‚¨ ModelEvaluationResult å¯¹è±¡
            eval_successful = False

            try:
                raw_report_from_evalscope = run_task(task_cfg=task_cfg)
                current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Evalscope run_task completed.")
                eval_successful = True

                if isinstance(raw_report_from_evalscope, dict):
                    for ds_name_key, report_obj in raw_report_from_evalscope.items():
                        evalscope_final_report[ds_name_key] = serialize_evalscope_report(report_obj)
                    current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Evalscope report processed and serialized.")
                else:
                    current_app.logger.error(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Evalscope run_task did not return a dictionary as expected. Got: {type(raw_report_from_evalscope)}")
                    evalscope_final_report = {"error": "Evalscope did not return a dictionary.", "raw_output": str(raw_report_from_evalscope)}
                    eval_successful = False # æ ‡è®°evalscopeå¤„ç†æŠ¥å‘Šéƒ¨åˆ†å¤±è´¥

                # å¦‚æœEvalscopeæ‰§è¡ŒæˆåŠŸå¹¶ä¸”æˆ‘ä»¬è·å¾—äº†æŠ¥å‘Šå­—å…¸ï¼Œå°è¯•è§£æè¯¦ç»†ç»“æœ
                if eval_successful and isinstance(evalscope_final_report, dict) and not evalscope_final_report.get("error"):
                    current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Attempting to parse detailed review files.")
                    # ç¡®ä¿base_output_diræ˜¯ç»å¯¹è·¯å¾„
                    if not os.path.isabs(base_output_dir):
                        base_output_dir = os.path.abspath(base_output_dir)
                    t_base_output_dir = base_output_dir
                    for k in os.listdir(base_output_dir):
                        t_base_output_dir = os.path.join(base_output_dir, k)
                    
                    # fix: model_to_evaluate.model_identifierå¯èƒ½æ˜¯deepseek/deepseek-r1-0528-qwen3-8bè¿™ç§æ ¼å¼ï¼Œéœ€è¦åšä¸ªå¤„ç†
                    t_model_identifier = model_to_evaluate.model_identifier.split('/')[-1]
                    reviews_base_path = os.path.join(t_base_output_dir, OUTPUTS_STRUCTURE_REVIEWS_DIR, t_model_identifier)
                    if os.path.isdir(reviews_base_path):
                        for review_filename_in_dir in os.listdir(reviews_base_path):
                            review_file_path = os.path.join(reviews_base_path, review_filename_in_dir)
                            
                            if os.path.isfile(review_file_path) and review_filename_in_dir.endswith('.jsonl'):
                                current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Processing review file: {review_file_path}")
                                
                                filename_stem = review_filename_in_dir[:-6] # Remove .jsonl
                                
                                # æ ¹æ®filename_stemæŸ¥æ‰¾å¯¹åº”çš„dataset
                                corresponding_dataset = None
                                for assoc in eval_dataset_associations:
                                    dataset = Dataset.query.get(assoc.dataset_id)
                                    if dataset:
                                        if dataset.dataset_type == 'ç³»ç»Ÿ':
                                            # ç³»ç»Ÿæ•°æ®é›†ç›´æ¥æ¯”è¾ƒåç§°
                                            if dataset.name in filename_stem:
                                                corresponding_dataset = dataset
                                                break
                                        elif dataset.dataset_type == 'è‡ªå»º':
                                            # è‡ªå»ºæ•°æ®é›†æ¯”è¾ƒæ–‡ä»¶åï¼ˆå»æ‰æ‰©å±•ååçš„éƒ¨åˆ†ï¼‰
                                            if dataset.download_url:
                                                dataset_filename = os.path.splitext(os.path.basename(dataset.download_url))[0]
                                                prefix = ''
                                                if dataset.format.lower() == 'custom':
                                                    prefix = f'custom_dataset_{dataset.id}'
                                                elif dataset.format.lower() == 'qa':
                                                    prefix = 'general_qa'
                                                elif dataset.format.lower() == 'mcq':
                                                    prefix = 'general_mcq'
                                                else:
                                                    prefix = f'custom_dataset_{dataset.id}'
                                                if f'{prefix}_{dataset_filename}' == filename_stem:
                                                    corresponding_dataset = dataset
                                                    break
                                
                                if not corresponding_dataset:
                                    current_app.logger.warning(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] æ— æ³•æ‰¾åˆ°filename_stem '{filename_stem}' å¯¹åº”çš„æ•°æ®é›†ï¼Œè·³è¿‡è¯¥æ–‡ä»¶")
                                    continue

                                try:
                                    origin_df = pd.read_json(review_file_path, lines=True)
                                    for _, item in origin_df.iterrows():
                                        raw_input = item.get('raw_input', '')
                                        raw_pred_answer = ''
                                        choices = item.get('choices', [])
                                        if choices and isinstance(choices, list) and len(choices) > 0:
                                            choice = choices[0]
                                            if isinstance(choice, dict) and 'message' in choice and isinstance(choice['message'], dict):
                                                raw_pred_answer = choice['message'].get('content', '')
                                            elif isinstance(choice, dict) and 'content' in choice:
                                                 raw_pred_answer = choice.get('content', '')
                                        
                                        review_data = {}
                                        if choices and isinstance(choices, list) and len(choices) > 0 and isinstance(choices[0], dict):
                                            review_data = choices[0].get('review', {})

                                        parsed_gold_answer = review_data.get('gold', '')
                                        parsed_pred_answer_for_feedback = review_data.get('pred', '') 
                                        score = review_data.get('result')

                                        # å¤„ç†ä¸åŒæ ¼å¼çš„result
                                        try:
                                            if score is not None:
                                                if isinstance(score, dict):
                                                    # å¤„ç†å¤åˆç»“æœæ ¼å¼: {"intent_result": true, "slots_result": {"miss_count": 1, "correct_count": 1, "fail_count": 0}}
                                                    if 'intent_result' in score and 'slots_result' in score:
                                                        intent_result = score.get('intent_result', False)
                                                        slots_result = score.get('slots_result', {})
                                                        
                                                        # è®¡ç®—slotçš„F1åˆ†æ•°
                                                        correct_count = slots_result.get('correct_count', 0)
                                                        miss_count = slots_result.get('miss_count', 0)
                                                        fail_count = slots_result.get('fail_count', 0)
                                                        
                                                        total_predicted = correct_count + fail_count
                                                        total_actual = correct_count + miss_count
                                                        
                                                        if total_predicted > 0 and total_actual > 0:
                                                            precision = correct_count / total_predicted
                                                            recall = correct_count / total_actual
                                                            if precision + recall > 0:
                                                                slot_f1 = 2 * precision * recall / (precision + recall)
                                                            else:
                                                                slot_f1 = 0.0
                                                        else:
                                                            slot_f1 = 0.0
                                                        
                                                        if correct_count + miss_count + fail_count == 0:
                                                            slot_f1 = 1.0
                                                        
                                                        score = float(intent_result) * 0.5 + 0.5 * slot_f1
                                                        current_app.logger.debug(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] å¤åˆç»“æœè®¡ç®—: intent={intent_result}, slot_f1={slot_f1:.4f}, final_score={score:.4f}")
                                                    else:
                                                        # å…¶ä»–å­—å…¸æ ¼å¼ï¼Œå°è¯•è½¬æ¢ä¸ºfloat
                                                        score = float(score)
                                                else:
                                                    # åŸæœ‰çš„ç®€å•æ•°å€¼æ ¼å¼
                                                    score = float(score)
                                        except (ValueError, TypeError) as e:
                                            current_app.logger.warning(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Could not parse score '{score}' for an item in {review_file_path}. Error: {str(e)}. Setting to None.")
                                            score = None

                                        result_entry = ModelEvaluationResult(
                                            evaluation_id=evaluation.id,
                                            dataset_id=corresponding_dataset.id,  # ä½¿ç”¨corresponding_dataset
                                            question=str(raw_input),
                                            model_answer=str(raw_pred_answer),
                                            reference_answer=str(parsed_gold_answer),
                                            score=score,
                                            feedback=str(parsed_pred_answer_for_feedback) 
                                        )
                                        detailed_results_to_save.append(result_entry)
                                    current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Processed {len(detailed_results_to_save)} items from {review_file_path} (total for this file: {len(origin_df)})")
                                except Exception as df_exc:
                                    current_app.logger.error(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Error processing review file {review_file_path}: {str(df_exc)}", exc_info=True)
                            else:
                                current_app.logger.debug(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Skipping non-JSONL file or directory in reviews folder: {review_filename_in_dir}")
                    else:
                        current_app.logger.warning(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Reviews directory not found: {reviews_base_path}. Skipping detailed results parsing.")
                
                # ä¿å­˜è¯¦ç»†ç»“æœåˆ°æ•°æ®åº“
                if detailed_results_to_save:
                    db.session.bulk_save_objects(detailed_results_to_save)
                    current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Saved {len(detailed_results_to_save)} detailed judge results to database.")

                evaluation.result_summary = evalscope_final_report
                evaluation.status = 'completed' if eval_successful else 'failed' # å¦‚æœevalscopeæ‰§è¡Œæœ¬èº«å°±å¤±è´¥äº†ï¼Œåˆ™æœ€ç»ˆçŠ¶æ€ä¸ºfailed
                if eval_successful and not evalscope_final_report.get("error"): # ä»…å½“evalscopeæˆåŠŸä¸”æŠ¥å‘Šæœ‰æ•ˆæ—¶æ ‡è®°å®Œæˆ
                     evaluation.status = 'completed'
                else: # å…¶ä»–æƒ…å†µéƒ½ç®—å¤±è´¥
                    evaluation.status = 'failed'
                    if not evalscope_final_report.get("error") and eval_successful: # evalscopeæ‰§è¡Œå®Œä½†æŠ¥å‘Šä¸ºç©ºæˆ–édict
                         evalscope_final_report["error_detail"] = "Evalscope run task finished but report was invalid or empty."
                    evaluation.result_summary = evalscope_final_report


                evaluation.completed_at = get_beijing_time()
                db.session.commit() # æäº¤æ‰€æœ‰æ›´æ”¹ï¼ŒåŒ…æ‹¬çŠ¶æ€ã€æ‘˜è¦å’Œè¯¦ç»†ç»“æœ
                current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] è¯„ä¼°ä»»åŠ¡å¤„ç†å®Œæ¯•ï¼ŒçŠ¶æ€: {evaluation.status}ã€‚Summary: {json.dumps(evalscope_final_report, indent=2)}")
                
                # æ¸…ç†è¯„ä¼°ç¼“å­˜
                EvaluationService._clear_evaluation_cache(evaluation_id)

            except Exception as es_exc:
                current_app.logger.error(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Error during evalscope execution or result processing: {str(es_exc)}", exc_info=True)
                evaluation.status = 'failed'
                evaluation.result_summary = {"error": f"Evalscope execution/processing failed: {str(es_exc)}"}
                db.session.commit() # ç¡®ä¿å³ä½¿å‘ç”Ÿå¼‚å¸¸ä¹Ÿæäº¤çŠ¶æ€
                
                # æ¸…ç†è¯„ä¼°ç¼“å­˜
                EvaluationService._clear_evaluation_cache(evaluation_id)
            
            finally:
                if os.path.isdir(base_output_dir):
                    try:
                        # shutil.rmtree(base_output_dir)
                        current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Successfully cleaned up evalscope output directory: {base_output_dir}")
                    except Exception as cleanup_exc:
                        current_app.logger.error(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Failed to clean up evalscope output directory {base_output_dir}: {str(cleanup_exc)}")
                else:
                    current_app.logger.warning(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] Evalscope output directory not found for cleanup: {base_output_dir}")
            current_app.logger.info(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] æ‰§è¡Œçº¿ç¨‹ç»“æŸã€‚")

    @staticmethod
    def get_evaluation_by_id(evaluation_id: int, user_id: int) -> Optional[ModelEvaluation]:
        evaluation = ModelEvaluation.query.get(evaluation_id)
        if not evaluation or evaluation.user_id != user_id:
            return None
        return evaluation
    
    @staticmethod
    def get_evaluations_for_user(user_id: int, page: int = 1, per_page: int = 10) -> Tuple[List[ModelEvaluation], int]:
        query = ModelEvaluation.query.filter_by(user_id=user_id).order_by(ModelEvaluation.created_at.desc())
        total = query.count()
        evaluations = query.paginate(page=page, per_page=per_page, error_out=False).items
        return evaluations, total
    
    @staticmethod
    def get_evaluation_results(
        evaluation_id: int, 
        user_id: int, 
        page: int = 1, 
        per_page: int = 10,
        search_query: Optional[str] = None,  # æœç´¢æŸ¥è¯¢å‚æ•°
        min_score: Optional[float] = None,   # æœ€å°åˆ†æ•°ç­›é€‰
        max_score: Optional[float] = None    # æœ€å¤§åˆ†æ•°ç­›é€‰
    ) -> Tuple[List[ModelEvaluationResult], int]:
        evaluation = ModelEvaluation.query.get(evaluation_id)
        if not evaluation or evaluation.user_id != user_id:
            return [], 0
        
        query = ModelEvaluationResult.query.filter_by(evaluation_id=evaluation_id)
        
        # å¦‚æœæä¾›äº†æœç´¢æŸ¥è¯¢ï¼Œåˆ™æ·»åŠ æ¨¡ç³Šæœç´¢æ¡ä»¶
        if search_query:
            query = query.filter(ModelEvaluationResult.question.ilike(f"%{search_query}%"))
        
        # æ·»åŠ åˆ†æ•°èŒƒå›´ç­›é€‰æ¡ä»¶
        if min_score is not None:
            query = query.filter(ModelEvaluationResult.score >= min_score)
        if max_score is not None:
            query = query.filter(ModelEvaluationResult.score <= max_score)
            
        query = query.order_by(ModelEvaluationResult.id.asc())
        
        total = query.count()
        results = query.paginate(page=page, per_page=per_page, error_out=False).items
        # ä¸ºæ¯ä¸ªç»“æœæ·»åŠ userPrompt
        for result in results:
            try:
                result.user_prompt = EvaluationService._get_user_prompt_for_result(result)
            except Exception as e:
                current_app.logger.error(f"è·å–ç»“æœ {result.id} çš„userPromptå¤±è´¥: {str(e)}")
                result.user_prompt = "æ— æ³•è·å–ç”¨æˆ·æç¤º"
        
        current_app.logger.info(f"[è¯„ä¼°ç»“æœæŸ¥è¯¢] EvalID: {evaluation_id}, UserID: {user_id}, Page: {page}, Search: '{search_query}', ScoreRange: [{min_score}, {max_score}], Found: {len(results)}, Total: {total}")
        return results, total

    @staticmethod
    def export_evaluation_results_to_excel(
        evaluation_id: int, 
        user_id: int, 
        search_query: Optional[str] = None,
        min_score: Optional[float] = None,
        max_score: Optional[float] = None
    ) -> Optional[bytes]:
        """
        å¯¼å‡ºè¯„ä¼°ç»“æœä¸ºExcelæ–‡ä»¶
        
        Args:
            evaluation_id: è¯„ä¼°ID
            user_id: ç”¨æˆ·ID
            search_query: æœç´¢æŸ¥è¯¢
            min_score: æœ€å°åˆ†æ•°
            max_score: æœ€å¤§åˆ†æ•°
            
        Returns:
            bytes: Excelæ–‡ä»¶çš„äºŒè¿›åˆ¶æ•°æ®ï¼Œå¤±è´¥è¿”å›None
        """
        try:
            # éªŒè¯æƒé™
            evaluation = ModelEvaluation.query.get(evaluation_id)
            if not evaluation or evaluation.user_id != user_id:
                return None
            
            # æ„å»ºæŸ¥è¯¢
            query = ModelEvaluationResult.query.filter_by(evaluation_id=evaluation_id)
            
            # åº”ç”¨ç­›é€‰æ¡ä»¶
            if search_query:
                query = query.filter(ModelEvaluationResult.question.ilike(f"%{search_query}%"))
            if min_score is not None:
                query = query.filter(ModelEvaluationResult.score >= min_score)
            if max_score is not None:
                query = query.filter(ModelEvaluationResult.score <= max_score)
                
            query = query.order_by(ModelEvaluationResult.id.asc())
            
            # è·å–æ€»æ•°ä½†ä¸åŠ è½½æ•°æ®
            total_count = query.count()
            if total_count == 0:
                return None
            
            current_app.logger.info(f"å¼€å§‹å¯¼å‡ºè¯„ä¼°ç»“æœï¼Œæ€»æ•°: {total_count}")
            
            # åˆ›å»ºExcelæ–‡ä»¶
            from io import BytesIO
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # åˆ†æ‰¹å¤„ç†æ•°æ®ï¼Œé¿å…OOM
                batch_size = 100  # æ¯æ‰¹å¤„ç†1000æ¡è®°å½•
                all_data = []
                
                # ä½¿ç”¨yield_perè¿›è¡Œæµå¼æŸ¥è¯¢
                for batch_start in range(0, total_count, batch_size):
                    current_app.logger.info(f"å¤„ç†æ‰¹æ¬¡: {batch_start}-{min(batch_start + batch_size, total_count)}")
                    
                    # åˆ†é¡µæŸ¥è¯¢å½“å‰æ‰¹æ¬¡
                    batch_results = query.offset(batch_start).limit(batch_size).all()
                    
                    # å¤„ç†å½“å‰æ‰¹æ¬¡æ•°æ®
                    for idx, result in enumerate(batch_results, start=batch_start + 1):
                        # ä½¿ç”¨ç»Ÿä¸€çš„user_promptè·å–æ–¹æ³•
                        try:
                            formatted_question = EvaluationService._get_user_prompt_for_result(result)
                        except Exception as e:
                            current_app.logger.error(f"è·å–ç»“æœ {result.id} çš„userPromptå¤±è´¥: {str(e)}")
                            formatted_question = result.question
                        
                        all_data.append({
                            'åºå·': idx,
                            'é—®é¢˜': formatted_question,
                            'æ¨¡å‹å›ç­”': result.model_answer,
                            'å‚è€ƒç­”æ¡ˆ': result.reference_answer or 'æ— ',
                            'å¾—åˆ†': result.score if result.score is not None else 'æ— è¯„åˆ†',
                            'æ•°æ®é›†': result.dataset.name if result.dataset else 'æœªçŸ¥æ•°æ®é›†'
                        })
                    
                    # æ¸…ç†å½“å‰æ‰¹æ¬¡ä»¥é‡Šæ”¾å†…å­˜
                    del batch_results
                
                # åˆ›å»ºDataFrame
                df = pd.DataFrame(all_data)
                current_app.logger.info(f"åˆ›å»ºDataFrameå®Œæˆï¼Œå…± {len(df)} è¡Œ")
                
                # å†™å…¥Excel
                df.to_excel(writer, sheet_name='è¯„ä¼°ç»“æœ', index=False)
                
                # è·å–å·¥ä½œè¡¨å¯¹è±¡è¿›è¡Œæ ¼å¼åŒ–
                worksheet = writer.sheets['è¯„ä¼°ç»“æœ']
                
                # è°ƒæ•´åˆ—å®½
                column_widths = {
                    'A': 8,   # åºå·
                    'B': 50,  # é—®é¢˜
                    'C': 50,  # æ¨¡å‹å›ç­”
                    'D': 30,  # å‚è€ƒç­”æ¡ˆ
                    'E': 10,  # å¾—åˆ†
                    'F': 20   # æ•°æ®é›†
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # è®¾ç½®è¡¨å¤´æ ·å¼
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # è®¾ç½®æ•°æ®è¡Œæ ·å¼
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # æ¸…ç†DataFrameä»¥é‡Šæ”¾å†…å­˜
                del df
                del all_data
            
            output.seek(0)
            file_data = output.getvalue()
            current_app.logger.info(f"Excelå¯¼å‡ºå®Œæˆï¼Œæ–‡ä»¶å¤§å°: {len(file_data)} bytes")
            return file_data
            
        except Exception as e:
            current_app.logger.error(f"å¯¼å‡ºExcelå¤±è´¥: {str(e)}", exc_info=True)
            return None 

    @staticmethod
    def _get_user_prompt_for_result(result: 'ModelEvaluationResult') -> str:
        """ä¸ºå•ä¸ªè¯„ä¼°ç»“æœè·å–æ ¼å¼åŒ–çš„userPrompt"""
        try:
            # è§£æquestionå­—æ®µè·å–åŸå§‹è¾“å…¥æ•°æ®
            import ast
            try:
                raw_input_data = json.loads(result.question)
            except (json.JSONDecodeError, TypeError):
                # å¦‚æœJSONè§£æå¤±è´¥ï¼Œå°è¯•ä½¿ç”¨ast.literal_eval
                try:
                    raw_input_data = ast.literal_eval(result.question)
                except (ValueError, SyntaxError):
                    # å¦‚æœéƒ½å¤±è´¥ï¼Œç›´æ¥è¿”å›åŸå§‹é—®é¢˜
                    return result.question
            
            # é¦–å…ˆå°è¯•ä½¿ç”¨adapterç”Ÿæˆå®Œæ•´çš„prompt
            # é€šè¿‡result.datasetè·å–æ•°æ®é›†ä¿¡æ¯ï¼Œç„¶åä½¿ç”¨benchmark_name
            dataset = result.dataset
            if not dataset:
                # å¦‚æœæ²¡æœ‰datasetå…³ç³»ï¼Œå›é€€åˆ°æ ¼å¼åŒ–é€»è¾‘
                return EvaluationService._format_prompt_from_raw_data(raw_input_data)
            
            benchmark_name = f'custom_dataset_{dataset.id}' if dataset.format.lower() == 'custom' else dataset.name
            adapter = EvaluationService.get_adapter_for_dataset(dataset.id)
            if adapter:
                try:
                    # è°ƒç”¨adapterçš„gen_promptæ–¹æ³•ï¼Œä¼ é€’æ­£ç¡®çš„å‚æ•°
                    prompt_data = adapter.gen_prompt(raw_input_data, benchmark_name, []) 
                    # æå–promptå­—æ®µ
                    if isinstance(prompt_data, dict):
                        adapter_prompt = prompt_data.get('data', []) or prompt_data.get('user_prompt', '')
                        if adapter_prompt:
                            # å¤„ç†ä¸åŒç±»å‹çš„adapter_prompt
                            if isinstance(adapter_prompt, list):
                                # å¦‚æœæ˜¯åˆ—è¡¨ï¼Œæ‹¼æ¥æˆå­—ç¬¦ä¸²
                                formatted_prompt = '\n'.join(str(item) for item in adapter_prompt)
                            else:
                                # å¦‚æœæ˜¯å­—ç¬¦ä¸²ï¼Œç›´æ¥ä½¿ç”¨
                                formatted_prompt = str(adapter_prompt)
                            
                            # å°†\nè½¬æ¢ä¸ºçœŸæ­£çš„æ¢è¡Œç¬¦
                            formatted_prompt = formatted_prompt.replace('\\n', '\n')
                            
                            return formatted_prompt
                except Exception as e:
                    current_app.logger.warning(f"ä½¿ç”¨adapterç”Ÿæˆpromptå¤±è´¥: {str(e)}")
            
            # å¦‚æœadapteræ–¹æ³•å¤±è´¥ï¼Œä½¿ç”¨è‡ªå®šä¹‰æ ¼å¼åŒ–é€»è¾‘
            return EvaluationService._format_prompt_from_raw_data(raw_input_data)
                
        except Exception as e:
            current_app.logger.error(f"è§£æç»“æœuserPromptæ—¶å‡ºé”™: {str(e)}", exc_info=True)
            return "è§£æé”™è¯¯"

    @staticmethod
    def _format_prompt_from_raw_data(raw_input_data: dict) -> str:
        """ä»åŸå§‹æ•°æ®æ ¼å¼åŒ–å®Œæ•´çš„promptæ˜¾ç¤º"""
        try:
            formatted_parts = []
            
            # å¤„ç†å†å²å¯¹è¯
            history = raw_input_data.get('history') or raw_input_data.get('history', [])
            if history and isinstance(history, list):
                for turn_idx, turn in enumerate(history):
                    if isinstance(turn, dict):
                        if turn.get('user'):
                            formatted_parts.append(f"ğŸ‘¤ ç”¨æˆ·: {turn['user']}")
                        if turn.get('assistant'):
                            formatted_parts.append(f"ğŸ¤– åŠ©æ‰‹: {turn['assistant']}")
                    elif isinstance(turn, list) and len(turn) >= 2:
                        # å¤„ç† [user, assistant] æ ¼å¼
                        formatted_parts.append(f"ğŸ‘¤ ç”¨æˆ·: {turn[0]}")
                        formatted_parts.append(f"ğŸ¤– åŠ©æ‰‹: {turn[1]}")
                
                # å¦‚æœæœ‰å†å²å¯¹è¯ï¼Œæ·»åŠ åˆ†éš”ç¬¦
                if formatted_parts:
                    formatted_parts.append("â”€" * 50)
            
            # å¤„ç†å½“å‰é—®é¢˜ - å°è¯•å¤šä¸ªå¯èƒ½çš„å­—æ®µå
            current_question = None
            question_fields = ['question', 'user', 'query', 'prompt', 'input']
            
            for field in question_fields:
                if field in raw_input_data:
                    value = raw_input_data[field]
                    if isinstance(value, str) and value.strip():
                        current_question = value.strip()
                        break
            
            if current_question:
                formatted_parts.append(f"ğŸ‘¤ ç”¨æˆ·: {current_question}")
            
            # å¤„ç†é€‰æ‹©é¢˜é€‰é¡¹ï¼ˆå¦‚æœæ˜¯MCQæ ¼å¼ï¼‰
            if 'A' in raw_input_data or 'choices' in raw_input_data:
                choices_text = EvaluationService._format_choices(raw_input_data)
                if choices_text:
                    formatted_parts.append(choices_text)
            
            # å¤„ç†ç³»ç»Ÿæç¤ºï¼ˆå¦‚æœæœ‰ï¼‰
            if 'system' in raw_input_data and raw_input_data['system']:
                system_prompt = raw_input_data['system']
                formatted_parts.insert(0, f"ğŸ”§ ç³»ç»Ÿ: {system_prompt}")
                formatted_parts.insert(1, "â”€" * 50)
            
            if formatted_parts:
                return "\n".join(formatted_parts)
            else:
                # å¦‚æœæ‰€æœ‰å­—æ®µéƒ½ä¸ºç©ºï¼Œè¿”å›åŸå§‹æ•°æ®çš„å­—ç¬¦ä¸²è¡¨ç¤º
                return str(raw_input_data)
                
        except Exception as e:
            current_app.logger.error(f"æ ¼å¼åŒ–promptæ—¶å‡ºé”™: {str(e)}")
            return str(raw_input_data)

    @staticmethod
    def _format_choices(raw_input_data: dict) -> str:
        """æ ¼å¼åŒ–é€‰æ‹©é¢˜é€‰é¡¹"""
        try:
            choices_parts = []
            
            # æ–¹å¼1: å¤„ç† A, B, C, D æ ¼å¼çš„é€‰é¡¹
            option_keys = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
            found_options = []
            
            for key in option_keys:
                if key in raw_input_data and raw_input_data[key]:
                    found_options.append(f"{key}. {raw_input_data[key]}")
            
            if found_options:
                choices_parts.append("ğŸ“‹ é€‰é¡¹:")
                choices_parts.extend(found_options)
            
            # æ–¹å¼2: å¤„ç† choices æ•°ç»„æ ¼å¼
            elif 'choices' in raw_input_data:
                choices = raw_input_data['choices']
                if isinstance(choices, list):
                    choices_parts.append("ğŸ“‹ é€‰é¡¹:")
                    for idx, choice in enumerate(choices):
                        letter = chr(ord('A') + idx)
                        choices_parts.append(f"{letter}. {choice}")
            
            return "\n".join(choices_parts) if choices_parts else ""
            
        except Exception as e:
            current_app.logger.error(f"æ ¼å¼åŒ–é€‰é¡¹æ—¶å‡ºé”™: {str(e)}")
            return ""

    @staticmethod
    def get_adapter_for_dataset(dataset_id: int):
        """æ ¹æ®æ•°æ®é›†åç§°è·å–å¯¹åº”çš„adapterå®ä¾‹"""
        try:
            dataset = Dataset.query.get(dataset_id)
            if not dataset:
                return None
            dataset_name = f'custom_dataset_{dataset.id}' if dataset.format.lower() == 'custom' else dataset.name
            # å¯¼å…¥BENCHMARK_MAPPINGS
            from evalscope.benchmarks.benchmark import BENCHMARK_MAPPINGS
            # åŠ¨æ€æ³¨å†Œè‡ªå®šä¹‰æ•°æ®é›†åŸºå‡†æµ‹è¯•[é‡å¯åå†…å­˜æ•°æ®ä¼šä¸¢å¤±ï¼Œæ‰€ä»¥åŠ¨æ€æ³¨å†Œä¸‹]
            from app.adapter.custom_dataset_adapter import register_custom_dataset_benchmark
            register_custom_dataset_benchmark(dataset.id)
            # ç»Ÿä¸€é€šè¿‡BENCHMARK_MAPPINGSè·å–adapter
            if dataset_name in BENCHMARK_MAPPINGS:
                benchmark_meta = BENCHMARK_MAPPINGS[dataset_name]
                adapter_class = benchmark_meta.data_adapter
                if dataset.format.lower() == 'custom':
                    return adapter_class(**benchmark_meta.to_dict(), template_content=dataset.jinja2_template)
                else:
                    return adapter_class(**benchmark_meta.to_dict())
            return None
            
        except Exception as e:
            current_app.logger.error(f"è·å–æ•°æ®é›† {dataset_id} çš„adapterå¤±è´¥: {str(e)}")
            raise e

    @staticmethod
    def get_evaluation_progress(evaluation_id: int, user_id: int) -> Dict[str, Any]:
        """
        è·å–è¯„ä¼°ä»»åŠ¡çš„è¿›åº¦ä¿¡æ¯
        
        Args:
            evaluation_id: è¯„ä¼°ID
            user_id: ç”¨æˆ·ID
            
        Returns:
            DictåŒ…å«è¿›åº¦ä¿¡æ¯ï¼štotal_prompts, completed_prompts, progress_percentage, status
        """
        try:
            # éªŒè¯æƒé™
            evaluation = ModelEvaluation.query.get(evaluation_id)
            if not evaluation or evaluation.user_id != user_id:
                return {"error": "è¯„ä¼°ä¸å­˜åœ¨æˆ–æ‚¨æ— æƒè®¿é—®"}
            
            # å¦‚æœè¯„ä¼°å·²å®Œæˆæˆ–å¤±è´¥ï¼Œç›´æ¥è¿”å›çŠ¶æ€
            if evaluation.status in ['completed', 'failed']:
                return {
                    "status": evaluation.status,
                    "total_prompts": 0,
                    "completed_prompts": 0,
                    "progress_percentage": 100.0 if evaluation.status == 'completed' else 0.0
                }
            
            # è·å–è¯„ä¼°è¾“å‡ºç›®å½•
            evalscope_run_timestamp = evaluation.created_at.strftime('%Y%m%d_%H%M%S')
            base_output_dir = os.path.join(get_outputs_dir(), f'eval_{evaluation_id}_{evalscope_run_timestamp}')
            # æ£€æŸ¥è¾“å‡ºç›®å½•æ˜¯å¦å­˜åœ¨
            if not os.path.exists(base_output_dir):
                return {
                    "status": evaluation.status,
                    "total_prompts": 0,
                    "completed_prompts": 0,
                    "progress_percentage": 0.0
                }
            
            # è·å–æ¨¡å‹æ ‡è¯†ç¬¦
            model = AIModel.query.get(evaluation.model_id)
            if not model:
                return {"error": "è¢«è¯„ä¼°æ¨¡å‹ä¸å­˜åœ¨"}
            
            if not os.path.isabs(base_output_dir):
                base_output_dir = os.path.abspath(base_output_dir)
            t_base_output_dir = base_output_dir
            for k in os.listdir(base_output_dir):
                t_base_output_dir = os.path.join(base_output_dir, k)

            t_model_identifier = model.model_identifier.split('/')[-1]
            reviews_base_path = os.path.join(t_base_output_dir, OUTPUTS_STRUCTURE_REVIEWS_DIR, t_model_identifier)
            current_app.logger.info(f"reviews_base_path: {reviews_base_path}")
            # è®¡ç®—å·²å®Œæˆçš„promptæ•°é‡ï¼ˆé€šè¿‡reviewsç›®å½•ä¸­çš„jsonæ–‡ä»¶ï¼‰
            completed_prompts = EvaluationService._calculate_completed_prompts(reviews_base_path)
            
            # æ£€æŸ¥æ˜¯å¦éœ€è¦è®¡ç®—total_promptsï¼ˆåªåœ¨é¦–æ¬¡æˆ–completed_promptsä¸º0æ—¶è®¡ç®—ï¼‰
            total_prompts = 0
            progress_percentage = 0.0
            
            # æ£€æŸ¥ç¼“å­˜ä¸­æ˜¯å¦å·²æœ‰total_prompts
            if evaluation_id in _evaluation_total_prompts_cache:
                total_prompts = _evaluation_total_prompts_cache[evaluation_id]
            else:
                # å¦‚æœç¼“å­˜ä¸­æ²¡æœ‰ï¼Œåˆ™è®¡ç®—å¹¶ç¼“å­˜
                total_prompts = EvaluationService._calculate_total_prompts(evaluation_id)
                _evaluation_total_prompts_cache[evaluation_id] = total_prompts
            
            # è®¡ç®—è¿›åº¦ç™¾åˆ†æ¯”
            if total_prompts > 0:
                progress_percentage = min(100.0, (completed_prompts / total_prompts) * 100.0)
            
            return {
                "status": evaluation.status,
                "total_prompts": total_prompts,
                "completed_prompts": completed_prompts,
                "progress_percentage": round(progress_percentage, 2)
            }
            
        except Exception as e:
            current_app.logger.error(f"è·å–è¯„ä¼°è¿›åº¦å¤±è´¥: {str(e)}")
            return {"error": f"è·å–è¿›åº¦å¤±è´¥: {str(e)}"}
    
    @staticmethod
    def _calculate_total_prompts(evaluation_id: int) -> int:
        """
        è®¡ç®—è¯„ä¼°ä»»åŠ¡çš„æ€»promptæ•°é‡ï¼Œé€šè¿‡data_adapter.loadåŠ è½½æ•°æ®é›†
        """
        try:
            total_prompts = 0
            
            # è·å–è¯„ä¼°è®°å½•
            evaluation = ModelEvaluation.query.get(evaluation_id)
            if not evaluation:
                return 0
            # è·å–è¯„ä¼°å…³è”çš„æ•°æ®é›†
            eval_dataset_associations = ModelEvaluationDataset.query.filter_by(evaluation_id=evaluation_id).all()
            
            for assoc in eval_dataset_associations:
                dataset = Dataset.query.get(assoc.dataset_id)
                adapter = EvaluationService.get_adapter_for_dataset(dataset.id)
                data_dict = adapter.load(dataset_name_or_path=dataset.download_url)
                prompts = adapter.gen_prompts(data_dict=data_dict)

                limited_prompts = defaultdict(list)
                for subset_name, prompts_list in prompts.items():
                    limit = evaluation.limit or len(prompts_list)
                    for index, prompt in enumerate(prompts_list[:limit]):
                        limited_prompts[subset_name].append(prompt)
                    total_prompts += len(limited_prompts[subset_name])
            return total_prompts
            
        except Exception as e:
            current_app.logger.error(f"è®¡ç®—æ€»promptæ•°é‡å¤±è´¥: {str(e)}")
            return 0

    @staticmethod
    def _calculate_completed_prompts(reviews_base_path: str) -> int:
        """
        è®¡ç®—å·²å®Œæˆçš„promptæ•°é‡ï¼ˆé€šè¿‡reviewsç›®å½•ä¸­çš„jsonæ–‡ä»¶ï¼‰
        """
        try:
            if not os.path.exists(reviews_base_path):
                return 0
            
            completed_count = 0
            
            # éå†reviewsç›®å½•ä¸­çš„æ‰€æœ‰jsonlæ–‡ä»¶
            for review_filename in os.listdir(reviews_base_path):
                review_file_path = os.path.join(reviews_base_path, review_filename)
                if os.path.isfile(review_file_path) and review_filename.endswith('.jsonl'):
                    try:
                        # è®¡ç®—æ–‡ä»¶ä¸­çš„è¡Œæ•°ï¼ˆæ¯ä¸ªè¡Œä»£è¡¨ä¸€ä¸ªå®Œæˆçš„promptï¼‰
                        with open(review_file_path, 'r', encoding='utf-8') as f:
                            file_count = sum(1 for line in f if line.strip())
                            completed_count += file_count
                    except Exception as e:
                        current_app.logger.warning(f"è¯»å–reviewæ–‡ä»¶å¤±è´¥ {review_file_path}: {str(e)}")
                        continue
            
            return completed_count
            
        except Exception as e:
            current_app.logger.error(f"è®¡ç®—å·²å®Œæˆpromptæ•°é‡å¤±è´¥: {str(e)}")
            return 0 

    @staticmethod
    def _clear_evaluation_cache(evaluation_id: int) -> None:
        """
        æ¸…ç†è¯„ä¼°ä»»åŠ¡çš„ç¼“å­˜
        """
        global _evaluation_total_prompts_cache
        if evaluation_id in _evaluation_total_prompts_cache:
            del _evaluation_total_prompts_cache[evaluation_id]
            current_app.logger.debug(f"[è¯„ä¼°ä»»åŠ¡ {evaluation_id}] å·²æ¸…ç†ç¼“å­˜") 